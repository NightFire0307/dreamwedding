# -*- coding: utf-8 -*-
# Create By:Mr.Thunder
# Power By:Abnegate
# 2019/3/19 20:59

from openpyxl import load_workbook

from .models import Kepan

def Save_model(excel_file):
    wb = load_workbook(excel_file)
    sheet = wb.active
    max_row = sheet.max_row

    headers = ['order', 'name_man', 'name_wom',
               'status', 'created_time', 'note']
    lists = []

    for row in range(3, max_row + 1):
        context = {}
        for column in range(1, len(headers) + 1):
            key = headers[column - 1]
            if column != 5:
                context[key] = sheet.cell(row=row, column=column).value
            else:
                time = sheet.cell(row=row, column=column).value
                context[key] = '{}-{}-{}'.format(time.year, time.month, time.day)
        lists.append(context)

    sqllist = []
    for cell in lists:
        order = cell['order']
        name_man = cell['name_man']
        name_wom = cell['name_wom']
        if cell['status'] == '已完成':
            status = Kepan.STATUS_FINASH
        else:
            status = Kepan.STATUS_UNFINASHED
        created_time = cell['created_time']
        note = cell['note']
        sql = Kepan(order=order, name_man=name_man, name_wom=name_wom,
                    status=status, created_time=created_time, note=note)
        sqllist.append(sql)
    Kepan.objects.bulk_create(sqllist)

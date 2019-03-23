# -*- coding: utf-8 -*-
# Create By:Mr.Thunder
# Power By:Abnegate
# 2019/3/22 21:49

from openpyxl import load_workbook

from .models import Yeji, Shejishi, Department

def Save_yejimodel(excel_file):
    '''
    此处接收通过request传递过来的Excel文件

    注意：批量导入重复的订单号会报错
    '''
    wb = load_workbook(excel_file)
    sheet = wb.active
    max_row = sheet.max_row

    headers = ['order', 'name', 'jiaogao',
               'taoxi', 'jiaxuan', 'owner',
               'chujian_date', 'note', 'zhangshu',
               'department']
    lists = []

    for row in range(3, max_row + 1):
        context = {}
        for column in range(1, len(headers) + 1):
            key = headers[column - 1]
            context[key] = sheet.cell(row=row, column=column).value
        lists.append(context)

    sqllist = []
    for cell in lists:
        order = cell['order']
        name = cell['name']
        jiaogao = cell['jiaogao']
        taoxi = cell['taoxi']
        jiaxuan = cell['jiaxuan']
        owner = cell['owner']
        chujian_date = cell['chujian_date']
        note = cell['note']
        zhangshu = cell['zhangshu']
        department = str(cell['department'])
        sjsid = Shejishi.objects.get_or_create(name=owner)[0].id
        departmentid = Department.objects.get_or_create(name=department)[0].id
        sql = Yeji(order=order, name=name, jiaogao=jiaogao,
                   taoxi=taoxi, jiaxuan=jiaxuan, owner_id=sjsid, jiaogao_own_id=sjsid,
                   chujian_date=chujian_date, note=note, zhangshu=zhangshu,
                   department_id=departmentid)
        sqllist.append(sql)
    Yeji.objects.bulk_create(sqllist)


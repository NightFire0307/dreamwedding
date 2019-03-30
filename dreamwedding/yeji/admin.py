# -*- coding:utf-8 -*-
import os
from django.contrib import admin
from django.contrib.admin import helpers
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django.urls import path
from django.shortcuts import render
from daterange_filter.filter import DateRangeFilter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO

from .models import Shejishi, Yeji
from .views import download_yejiexcel

from kepan.forms import ExcelForms
from kepan.admin import is_excel
from .excel_save import Save_yejimodel

@admin.register(Shejishi)
class ShejishiAdmin(admin.ModelAdmin):
    list_display = ['name']

# @admin.register(Department)
# class DepartmentAdmin(admin.ModelAdmin):
#     list_display = ['name']

@admin.register(Yeji)
class YejiAdmin(admin.ModelAdmin):
    list_display = ['order', 'name', 'colored_jiaogao',
                    'taoxi', 'jiaxuan', 'owner',
                    'jiaogao_own', 'colored_chujian', 'zhangshu',
                    'note', 'department']

    list_filter = ('owner', 'jiaogao_own', 'department',
                   ('chujian_date', DateRangeFilter),
                   ('jiaogao', DateRangeFilter),)

    # list_editable = ['colored_chujian']

    save_on_top = True
    actions = ['export_excel']
    actions_on_top = True
    actions_on_bottom = True
    search_fields = ['order', 'jiaogao_own__name']

    change_list_template = 'admin/yeji/extras/change_list.html'

    def get_urls(self):
        urls = super(YejiAdmin, self).get_urls()
        custom_urls = [
            path('upload_yejiexcel/', self.admin_site.admin_view(self.upload_excel),
                 name='upload_yejiexcel'),
            path('download_yejiexcel', self.admin_site.admin_view(download_yejiexcel),
                 name='download_yejiexcel'),
        ]
        return custom_urls + urls

    def upload_excel(self, request):
        context = {
            'title': 'Upload Excel',
            'app_label': self.model._meta.app_label,
            'opts': self.model._meta,
            'has_change_permission': self.has_change_permission(request),
        }

        if request.method == 'POST':
            form = ExcelForms(request.POST, request.FILES)
            if form.is_valid():
                if is_excel(request.FILES['excel_file']):
                    Save_yejimodel(request.FILES['excel_file'])
                    return HttpResponseRedirect('..')
                else:
                    return HttpResponseRedirect('.')
        else:
            form = ExcelForms()

        context['form'] = form
        context['adminform'] = helpers.AdminForm(form,
                                                 list([(None, {'fields': form.base_fields})]),
                                                 {})
        return render(request, 'admin/yeji/extras/upload_excel.html', context)

    def export_excel(self, request, queryset):
        response = HttpResponse(content_type='applicationi/ms-excel')
        response['Content-Disposition'] = 'attachment;filename=OutPut.xlsx'

        wb = Workbook()
        ws = wb.active

        header = ['订单号', '客户姓名', '校稿时间',
                  '套系金额', '加选金额', '设计人员',
                  '出件时间', '备注', '张数', '订单所属部门']

        for column in range(1, len(header) + 1):
            cell_style = ws.cell(row=1, column=column, value=header[column - 1])
            cell_style.font = Font(name='宋体', size=12, bold=True, color="FFFFFF")
            cell_style.fill = PatternFill("solid", fgColor='8E8E8E')

        print(queryset)
        data_row = queryset.count() + 2
        for row in range(2, data_row):
            for column, data in enumerate(queryset, start=1):
                ws.cell(row=row, column=column, value=data.order)
                ws.cell(row=row, column=column+1, value=data.name)
                ws.cell(row=row, column=column+2, value=data.jiaogao).number_format = 'yyyy/mm/dd'
                ws.cell(row=row, column=column+3, value=data.taoxi)
                ws.cell(row=row, column=column+4, value=data.jiaxuan)
                ws.cell(row=row, column=column+5, value=data.owner.name)
                ws.cell(row=row, column=column+6, value=data.chujian_date)
                ws.cell(row=row, column=column+7, value=data.note)
                ws.cell(row=row, column=column+8, value=data.zhangshu)
                ws.cell(row=row, column=column+9, value=data.department)
        output = BytesIO()
        wb.save(output)
        # 初始化指针
        output.seek(0)
        response.write(output.getvalue())
        return response

    export_excel.short_description = '导出选中数据至Excel'
